import requests
import time
from datetime import datetime, timedelta
from openpyxl import load_workbook

# ─────────────────────────────────────────────
#  CONFIG
# ─────────────────────────────────────────────
TELEGRAM_TOKEN = "8430077568:AAEE2LBikDWtrx8j1iZvgIckXNlJl3xnGmA"
GROUP_CHAT_ID  = "-5118811032"
EXCEL_FILE     = "roster.xlsx"
SEND_HOUR      = 16   # 9 PM Maldives = 4 PM UTC
SEND_MINUTE    = 0

# ─────────────────────────────────────────────
#  STEP 1 — Parse roster from Excel
# ─────────────────────────────────────────────
def parse_roster(filepath):
    wb = load_workbook(filepath, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    roster = {}
    i = 0

    while i < len(rows):
        row = rows[i]
        month_val = row[2] if len(row) > 2 else None

        if month_val in ("APRIL", "MAY"):
            month_row = rows[i]
            date_row  = rows[i + 1]
            i += 3

            col_dates = {}
            current_month = None
            year = 2026

            for col_idx in range(len(date_row)):
                if col_idx < len(month_row) and month_row[col_idx] in ("APRIL", "MAY"):
                    current_month = 4 if month_row[col_idx] == "APRIL" else 5
                date_val = date_row[col_idx]
                if date_val and isinstance(date_val, str) and date_val[:-2].isdigit() and current_month:
                    col_dates[col_idx] = datetime(year, current_month, int(date_val[:-2]))

            while i < len(rows):
                staff_row = rows[i]
                if staff_row[0] is None and staff_row[1] is None:
                    break
                if staff_row[1] == "NAME":
                    break
                name = staff_row[1]
                if name:
                    name = name.strip()
                    for col_idx, date in col_dates.items():
                        shift = staff_row[col_idx] if col_idx < len(staff_row) else None
                        if shift is not None:
                            if date not in roster:
                                roster[date] = {}
                            roster[date][name] = str(shift).strip()
                i += 1
        else:
            i += 1

    return roster


# ─────────────────────────────────────────────
#  STEP 2 — Build the message
# ─────────────────────────────────────────────
def build_message(roster, target_date):
    day_name = target_date.strftime("%A")
    date_str = target_date.strftime("%-d %B")

    shifts  = {}
    offs    = []
    leaves  = []
    travels = []

    day_data = roster.get(target_date, {})
    if not day_data:
        return f"⚠️ No roster data found for {day_name}, {date_str}."

    for name, shift in sorted(day_data.items()):
        if shift == "OFF":
            offs.append(name)
        elif shift == "ANNUAL LEAVE":
            leaves.append(name)
        elif shift == "DUTY TRAVEL":
            travels.append(name)
        else:
            shifts.setdefault(shift, []).append(name)

    emoji_map = {
        "0600-1400": "🌅",
        "1200-2000": "🌤",
        "1400-2200": "🌙",
    }

    lines = [f"👋 Hey team! Tomorrow's duty ({day_name}, {date_str}):\n"]

    for shift_time in sorted(shifts.keys()):
        icon  = emoji_map.get(shift_time, "⏰")
        names = ", ".join(shifts[shift_time])
        lines.append(f"{icon} {shift_time} → {names}")

    if offs:
        lines.append(f"\n🏖 OFF → {', '.join(offs)}")
    if leaves:
        lines.append(f"✈️ Annual Leave → {', '.join(leaves)}")
    if travels:
        lines.append(f"🚗 Duty Travel → {', '.join(travels)}")

    lines.append("\nGood luck everyone! 💪")
    return "\n".join(lines)


# ─────────────────────────────────────────────
#  STEP 3 — Send to Telegram
# ─────────────────────────────────────────────
def send_to_telegram(message):
    print(f"📤 Sending to Telegram...")
    url     = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage"
    payload = {"chat_id": GROUP_CHAT_ID, "text": message}
    try:
        response = requests.post(url, json=payload, timeout=10)
        print(f"📡 Status: {response.status_code} | Body: {response.text}")
        result = response.json()
        if result.get("ok"):
            print(f"✅ Sent successfully!")
        else:
            print(f"❌ Failed: {result}")
    except Exception as e:
        print(f"❌ Exception: {e}")


# ─────────────────────────────────────────────
#  STEP 4 — Daily job
# ─────────────────────────────────────────────
def daily_job():
    now      = datetime.utcnow()
    tomorrow = now.replace(hour=0, minute=0, second=0, microsecond=0) + timedelta(days=1)

    print(f"\n⏰ Triggered! UTC now: {now.strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"📅 Looking for date: {tomorrow.strftime('%Y-%m-%d')}")

    try:
        roster = parse_roster(EXCEL_FILE)
        dates  = sorted(roster.keys())
        print(f"✅ Roster loaded: {len(dates)} days ({dates[0].date()} to {dates[-1].date()})")
        print(f"🔍 Is {tomorrow.date()} in roster? {tomorrow in roster}")

        message = build_message(roster, tomorrow)
        print(f"📋 Message:\n{message}\n")
        send_to_telegram(message)
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()


# ─────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────
if __name__ == "__main__":
    print(f"🤖 Bot started (UTC): {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"📅 Will send at {SEND_HOUR:02d}:{SEND_MINUTE:02d} UTC (9 PM Maldives)\n")

    sent_today = False

    while True:
        now = datetime.utcnow()

        if now.hour == 0 and now.minute == 0:
            sent_today = False

        if now.hour == SEND_HOUR and now.minute == SEND_MINUTE and not sent_today:
            print(f"🔔 Time to send!")
            daily_job()
            sent_today = True

        if now.second < 10:
            print(f"💓 {now.strftime('%H:%M')} UTC | Maldives: {(now + timedelta(hours=5)).strftime('%H:%M')}")

        time.sleep(10)
